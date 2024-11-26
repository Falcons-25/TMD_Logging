from openpyxl import load_workbook
import time

def average(a: list|tuple) -> float:
    return sum(a)/len(a)

def setup():
    global all_tstamps, all_thr, data, config
    all_tstamps = {}
    all_thr = {}
    with open("config.txt", 'r') as file:
        # read start point
        pass
    with open("TM_log.csv", 'r') as file:
        data = [x.strip() for x in file.readlines() if x.strip()]
    # print(data)
    config = data[0].strip().split(',')
    print(config)
    test_date = time.strftime("%d-%b-%Y", time.strptime(config[0].strip(), "%Y-%m-%d %H:%M:%S"))
    prop = config[1].strip()
    motor_kv = config[2].strip()
    esc = config[3].strip()
    batt = config[4].strip()
    config[:] = [test_date, prop, motor_kv, esc, batt]
    print(config)

def summarise_data():
    global final_tstamps, final_thr
    all_thr = []
    last_throttle = -1
    for line in data:
        if not line: continue
        # set start date and time
        values = line.split(',')
        if values[1].split()[0] in ["APC", "EMAX", "GemFan", "Generic"]: continue
        # timestamp, throttle, thrust, current, c2, volt, v2, rpm, discharge
        if values[0] not in all_tstamps.keys():
            all_tstamps[values[0]] = {"Throttle":[], "Thrust":[], "Current":[], "Current2":[], "Volt":[], "Volt2":[], "RPM":[], "Capacity":[]}
        all_tstamps[values[0]]["Throttle"].append(int(values[1]))
        all_tstamps[values[0]]["Thrust"].append(float(values[2]))
        all_tstamps[values[0]]["Current"].append(float(values[3]))
        all_tstamps[values[0]]["Current2"].append(float(values[4]))
        all_tstamps[values[0]]["Volt"].append(float(values[5]))
        all_tstamps[values[0]]["Volt2"].append(float(values[6]))
        all_tstamps[values[0]]["RPM"].append(int(values[7]))
        all_tstamps[values[0]]["Capacity"].append(int(values[8]))

        if int(values[1])==last_throttle:
            all_thr[-1][1].append(float(values[2]))
            all_thr[-1][2].append(float(values[3]))
            all_thr[-1][3].append(float(values[4]))
            all_thr[-1][4].append(float(values[5]))
            all_thr[-1][5].append(float(values[6]))
            all_thr[-1][6].append(int(values[7]))
            all_thr[-1][7].append(int(values[8]))
        else:
            last_throttle = int(values[1])
            all_thr.append([last_throttle, [float(values[2])], [float(values[3])], [float(values[4])], [float(values[5])], [float(values[6])], [int(values[7])], [int(values[8])]])

    # update dict with averaged values
    final_tstamps = {}
    for timestamp in all_tstamps:
        if timestamp not in final_tstamps.keys():
            final_tstamps[timestamp] = {"Throttle":[], "Thrust":[], "Current":[], "Current2":[], "Volt":[], "Volt2":[], "RPM":[], "Capacity":[]}
        final_tstamps[timestamp]["Throttle"] = int(average(all_tstamps[timestamp]["Throttle"]))
        final_tstamps[timestamp]["Thrust"] = average(all_tstamps[timestamp]["Thrust"])
        final_tstamps[timestamp]["Current"] = average(all_tstamps[timestamp]["Current"])
        final_tstamps[timestamp]["Current2"] = average(all_tstamps[timestamp]["Current2"])
        final_tstamps[timestamp]["Volt"] = average(all_tstamps[timestamp]["Volt"])
        final_tstamps[timestamp]["Volt2"] = average(all_tstamps[timestamp]["Volt2"])
        final_tstamps[timestamp]["RPM"] = int(average(all_tstamps[timestamp]["RPM"]))
        final_tstamps[timestamp]["Capacity"] = int(average(all_tstamps[timestamp]["Capacity"]))

    # update dict with averaged values for throttle
    final_thr = []
    for values in all_thr:
        final_thr.append([values[0], average(values[1]), average(values[2]), average(values[3]), average(values[4]), average(values[5]), int(average(values[6])), int(average(values[7]))])
    

def write_to_files():
    name = tm_type[0] + config[0]
    print("Sheet name:", name)
    wb1 = load_workbook("")
    # check for page
    if (name) in wb1:
        ws1 = wb1[name]
    else:
        ws1 = wb1.create_sheet(name)
    wb1.active = ws1
    start_row = ws1.max_row + 3
    print("(1) Start row:", start_row)
    
    ws1.cell(row=start_row+0, column=2).value = f"{config[2]}"
    print(f"{config[2]}")
    ws1.merge_cells(start_column=2, start_row=start_row, end_column=11, end_row=start_row)
    ws1.cell(row=start_row+1, column=2).value = f"({tm_type}) {config[0]}"
    print(f"({tm_type}) {config[0]}")
    ws1.merge_cells(start_column=2, start_row=start_row+1, end_column=11, end_row=start_row+1)
    ws1.cell(row=start_row+2, column=2).value = f"{config[1]}"
    print(f"{config[1]}")
    ws1.merge_cells(start_column=2, start_row=start_row+2, end_column=11, end_row=start_row+2)
    ws1.cell(row=start_row+3, column=2).value = f"{config[3]}"
    print(f"{config[3]}")
    ws1.merge_cells(start_column=2, start_row=start_row+3, end_column=11, end_row=start_row+3)
    ws1.cell(row=start_row+4, column=2).value = f"{config[4]}"
    print(f"{config[4]}")
    ws1.merge_cells(start_column=2, start_row=start_row+4, end_column=11, end_row=start_row+4)
    ws1.merge_cells(start_column=2, start_row=start_row+5, end_column=11, end_row=start_row+5)
    ws1.cell(row=start_row+6, column=2).value = "Timestamp"
    ws1.cell(row=start_row+6, column=3).value = "Throttle (%)"
    ws1.cell(row=start_row+6, column=4).value = "Thrust (gf)"
    ws1.cell(row=start_row+6, column=5).value = "Current (A)"
    ws1.cell(row=start_row+6, column=6).value = "Power (W)"
    ws1.cell(row=start_row+6, column=7).value = "Volt (V)"
    ws1.cell(row=start_row+6, column=8).value = "Current2 (A)"
    ws1.cell(row=start_row+6, column=9).value = "Volt2 (V)"
    ws1.cell(row=start_row+6, column=10).value = "RPM"
    ws1.cell(row=start_row+6, column=11).value = "Capacity (mAh)"
    print("Table headers configured.")
    start_row = start_row + 7
    wb1.save("")
    
    for i, tstamp in enumerate(final_tstamps.keys()):
        ws1.cell(row=start_row+i, column=2).value = tstamp
        ws1.cell(row=start_row+i, column=3).value = final_tstamps[tstamp]["Throttle"]
        ws1.cell(row=start_row+i, column=4).value = final_tstamps[tstamp]["Thrust"]
        ws1.cell(row=start_row+i, column=5).value = final_tstamps[tstamp]["Current"]
        ws1.cell(row=start_row+i, column=6).value = final_tstamps[tstamp]["Current"]*final_tstamps[tstamp]["Volt"]
        ws1.cell(row=start_row+i, column=7).value = final_tstamps[tstamp]["Volt"]
        ws1.cell(row=start_row+i, column=8).value = final_tstamps[tstamp]["Current2"]
        ws1.cell(row=start_row+i, column=9).value = final_tstamps[tstamp]["Volt2"]
        ws1.cell(row=start_row+i, column=10).value = final_tstamps[tstamp]["RPM"]
        ws1.cell(row=start_row+i, column=11).value = final_tstamps[tstamp]["Capacity"]
    # save file2
    wb1.save("")
    wb1.close()
    print("Timestamped data saved to file.")

    wb2 = load_workbook("")
    if name in wb2:
        ws2 = wb2[name]
    else:
        ws2 = wb2.create_sheet(name)
    wb2.active = ws2
    start_row = ws2.max_row + 3
    print("(2) Start row:", start_row)
    ws2.cell(row=start_row+0, column=2).value = f"{config[2]}"
    print(f"{config[2]}")
    ws2.merge_cells(start_column=2, start_row=start_row, end_column=10, end_row=start_row)
    ws2.cell(row=start_row+1, column=2).value = f"({tm_type}) {config[0]}"
    print(f"({tm_type}) {config[0]}")
    ws2.merge_cells(start_column=2, start_row=start_row+1, end_column=10, end_row=start_row+1)
    ws2.cell(row=start_row+2, column=2).value = f"{config[1]}"
    print(f"{config[1]}")
    ws2.merge_cells(start_column=2, start_row=start_row+2, end_column=10, end_row=start_row+2)
    ws2.cell(row=start_row+3, column=2).value = f"{config[3]}"
    print(f"{config[3]}")
    ws2.merge_cells(start_column=2, start_row=start_row+3, end_column=10, end_row=start_row+3)
    ws2.cell(row=start_row+4, column=2).value = f"{config[4]}"
    print(f"{config[4]}")
    ws2.merge_cells(start_column=2, start_row=start_row+4, end_column=10, end_row=start_row+4)
    ws2.merge_cells(start_column=2, start_row=start_row+5, end_column=10, end_row=start_row+5)
    ws2.cell(row=start_row+6, column=2).value = "Throttle (%)"
    ws2.cell(row=start_row+6, column=3).value = "Thrust (gf)"
    ws2.cell(row=start_row+6, column=4).value = "Current (A)"
    ws2.cell(row=start_row+6, column=5).value = "Power (W)"
    ws2.cell(row=start_row+6, column=6).value = "Volt (V)"
    ws2.cell(row=start_row+6, column=7).value = "Current2 (A)"
    ws2.cell(row=start_row+6, column=8).value = "Volt2 (V)"
    ws2.cell(row=start_row+6, column=9).value = "RPM"
    ws2.cell(row=start_row+6, column=10).value = "Capacity (mAh)"
    print("Table headers configured.")
    start_row = start_row + 7
    wb2.save("")

    for i, tstamp in enumerate(final_thr):
        ws2.cell(row=start_row+i, column=2).value = tstamp[0]
        ws2.cell(row=start_row+i, column=3).value = tstamp[1]
        ws2.cell(row=start_row+i, column=4).value = tstamp[2]
        ws2.cell(row=start_row+i, column=5).value = tstamp[2]*tstamp[4]
        ws2.cell(row=start_row+i, column=6).value = tstamp[4]
        ws2.cell(row=start_row+i, column=7).value = tstamp[3]
        ws2.cell(row=start_row+i, column=8).value = tstamp[5]
        ws2.cell(row=start_row+i, column=9).value = tstamp[6]
        ws2.cell(row=start_row+i, column=10).value = tstamp[7]
    # save file2
    wb2.save("")
    wb2.close()
    print("Throttle-stamped data saved to file.")

if __name__ == "__main__":
    tm_type = input("Static or Dynamic? ").strip()
    print("Setting up...")
    setup()
    print("Setup complete.")
    summarise_data()
    # write_to_files() tested, works
